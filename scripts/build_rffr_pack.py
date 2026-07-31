#!/usr/bin/env python3
"""Generate the RFFR (ISM SoA) content pack + public website catalogue.

Turns the department's RFFR Statement of Applicability workbook
(ISM June 2026) into two artefacts:

  1. checkpoint-content/rffr.json  -- the premium Checkpoint content pack
     (7 RFFR program-deed obligations + all applicable ISM controls),
     encrypted at build time like every other premium module. Not shipped
     in plaintext (checkpoint-content/ lives outside public/).

  2. src/data/rffr-controls.json   -- a PUBLIC catalogue the website
     renders statically. Carries only what is already public (the ISM is
     published by the ACSC; the SoA template by DEWR) -- control id, text,
     guideline, and the priority / E8 / classification / automation flags
     -- never the pack's proprietary implementation guidance.

Source of truth is the workbook, NOT this file -- re-run when a new ISM
quarter lands:

    python3 scripts/build_rffr_pack.py <path-to-workbook.xlsx>
"""
import json
import os
import sys
import warnings
from datetime import date

import openpyxl

warnings.filterwarnings("ignore")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = sys.argv[1] if len(sys.argv) > 1 else None
if not SRC:
    sys.exit("usage: build_rffr_pack.py <workbook.xlsx>")

wb = openpyxl.load_workbook(SRC, data_only=True)

# Guideline -> short category key (drives the SoA filter chips in app.js)
GUIDELINE_CAT = {
    "Guidelines for cyber security roles": "roles",
    "Guidelines for cyber security incidents": "incidents",
    "Guidelines for procurement and outsourcing": "procurement",
    "Guidelines for cyber security documentation": "documentation",
    "Guidelines for physical security": "physical",
    "Guidelines for personnel security": "personnel",
    "Guidelines for communications infrastructure": "comms-infra",
    "Guidelines for communications systems": "comms-systems",
    "Guidelines for enterprise mobility": "mobility",
    "Guidelines for evaluated products": "evaluated",
    "Guidelines for information technology equipment": "it-equipment",
    "Guidelines for media": "media",
    "Guidelines for system hardening": "hardening",
    "Guidelines for system management": "sys-mgmt",
    "Guidelines for security assurance": "assurance",
    "Guidelines for software development": "software-dev",
    "Guidelines for database systems": "database",
    "Guidelines for email": "email",
    "Guidelines for networking": "networking",
    "Guidelines for cryptography": "cryptography",
    "Guidelines for gateways": "gateways",
    "Guidelines for data transfers": "data-transfers",
}

# ---- Read ISM controls (data from row 3) ----
ws = wb["ISM June 2026"]
ism = []
for r in ws.iter_rows(min_row=3, values_only=True):
    if not r[3]:
        continue
    ism.append({
        "guideline": r[0], "section": r[1], "topic": r[2],
        "id": str(r[3]).strip(), "desc": str(r[4] or "").strip(),
        "rffrCore": r[5] == "Yes",
        "ml1": r[11] == "Yes", "ml2": r[12] == "Yes", "ml3": r[13] == "Yes",
        "nc": r[16] == "Yes", "os": r[17] == "Yes",
    })
ism_ids = {c["id"] for c in ism}

# ---- Read RFFR obligations (program deeds) ----
ws2 = wb["RFFR Obligations"]
deeds = []
for r in ws2.iter_rows(min_row=3, values_only=True):
    if not r[0]:
        continue
    num = "".join(ch for ch in str(r[0]) if ch.isdigit())
    deeds.append({"id": "RFFR-D" + num, "desc": str(r[1] or "").strip(),
                  "rffrCore": r[2] == "Yes"})

# ---- Microsoft posture-check -> ISM identifier automation map ----
# Curated for precision: each entry is an ISM control the tool's live
# Microsoft Graph signal genuinely speaks to. Suggest-only -- the tool
# never writes a status from these without practitioner confirmation
# (same contract as CHECK_E8 / CHECK_IS18).
CHECK_RFFR = {
    "mfa-all": ["ISM-1504", "ISM-1679", "ISM-1892", "ISM-1893", "ISM-1681", "ISM-0974"],
    "mfa-priv": ["ISM-1173"],
    "admins": ["ISM-1507", "ISM-1508", "ISM-1852"],
    "pim": ["ISM-1649", "ISM-1647"],
    "access-review": ["ISM-1647", "ISM-1509"],
    "patch": ["ISM-1876", "ISM-1690", "ISM-1691", "ISM-1692", "ISM-1698", "ISM-1699", "ISM-1701"],
    "macro": ["ISM-1671", "ISM-1488", "ISM-1672", "ISM-1673", "ISM-1674"],
    "wdac": ["ISM-0843", "ISM-1490", "ISM-1656", "ISM-1870", "ISM-1871"],
    "device": ["ISM-1400", "ISM-1482", "ISM-1195"],
    "labels": ["ISM-0271", "ISM-1187", "ISM-0272"],
    "encryption": ["ISM-1059", "ISM-1781"],
    "logging": ["ISM-1405", "ISM-1983", "ISM-1985", "ISM-1815"],
    "alerts": ["ISM-1417", "ISM-1288"],
    "backup": ["ISM-1511", "ISM-1810", "ISM-1811", "ISM-1515"],
}
# Prune any curated id absent from this ISM version (renumbering safety).
for k in list(CHECK_RFFR):
    CHECK_RFFR[k] = [i for i in CHECK_RFFR[k] if i in ism_ids]
    if not CHECK_RFFR[k]:
        del CHECK_RFFR[k]

# ISO 27001 / E8 anchor per check (mirrors store.js CHECK_CONTROLS) so the
# automatable ISM controls carry a cross-map that feeds the existing
# evidence-reuse engine.
ISO_ANCHOR = {
    "mfa-all": "ISO27001 A.5.15 · A.8.5 · E8.7", "mfa-priv": "ISO27001 A.8.2 · A.8.5 · E8.7",
    "admins": "ISO27001 A.8.2 · E8.5", "pim": "ISO27001 A.8.2 · A.5.18 · E8.5",
    "access-review": "ISO27001 A.5.18 · A.8.2", "patch": "ISO27001 A.8.8 · E8.2 · E8.6",
    "macro": "ISO27001 A.8.7 · E8.3", "wdac": "ISO27001 A.8.19 · A.8.7 · E8.1",
    "device": "ISO27001 A.8.1", "labels": "ISO27001 A.5.12 · A.5.13",
    "encryption": "ISO27001 A.8.24", "logging": "ISO27001 A.8.15",
    "alerts": "ISO27001 A.8.7 · A.8.16 · E8.1", "backup": "ISO27001 A.8.13 · E8.8",
}
ism_to_map, ism_to_check = {}, {}
for check, ids in CHECK_RFFR.items():
    for i in ids:
        if i not in ism_to_map:
            ism_to_map[i] = ISO_ANCHOR[check]
            ism_to_check[i] = check

# ---- Premium pack controls ----
pack_controls = []
for d in deeds:
    pack_controls.append({"code": d["id"], "t": d["desc"], "app": True,
                          "map": "ISO27001 A.6.1 · A.5.31", "cat": "deeds"})
for c in ism:
    pack_controls.append({
        "code": c["id"], "t": c["desc"], "app": True,
        "map": ism_to_map.get(c["id"], ""),
        "cat": GUIDELINE_CAT.get(c["guideline"], "other"),
    })

# ---- Guidance (proprietary): the automatable set gets a Microsoft hint ----
guidance = {}
for i, check in ism_to_check.items():
    guidance[i] = {
        "how": ('Checkpoint reads this from your Microsoft tenant via the "' + check +
                '" posture check — the live Graph signal proposes a status which a '
                'practitioner confirms before it is written to the SoA. No manual '
                'evidence gathering for the baseline.'),
        "evidence": ("Auto-captured scan result (dated, hashed) from Microsoft Graph, "
                     "plus the relevant Purview / Intune / Entra configuration export."),
    }

pack = {
    "moduleId": "rffr",
    "version": 1,
    "framework": {
        "id": "rffr",
        "name": "RFFR (ISM SoA)",
        "tag": "Cth Employment",
        "blurb": (
            "Right Fit For Risk — the DEWR cyber-security accreditation for Employment "
            "Services providers, delivered as one Statement of Applicability: the 7 "
            "program-deed obligations plus all " + str(len(ism)) + " Australian Government "
            "ISM (June 2026) controls applicable to Non-Classified and OFFICIAL: Sensitive "
            "information, cross-mapped to the ISO 27001 ISMS backbone and Essential Eight so "
            "the certification, the E8 uplift and the RFFR SoA are prepared once, not three "
            "times. RFFR Core Expectations are flagged for milestone prioritisation."
        ),
        "controls": pack_controls,
    },
    "guidance": guidance,
    "extra": {
        "checkRffr": CHECK_RFFR,
        "ismMeta": {c["id"]: {
            "g": c["guideline"], "sec": c["section"], "topic": c["topic"],
            "pri": c["rffrCore"], "ml1": c["ml1"], "ml2": c["ml2"], "ml3": c["ml3"],
            "nc": c["nc"], "os": c["os"],
        } for c in ism},
        "ismVersion": "June 2026",
    },
}
with open(os.path.join(ROOT, "checkpoint-content", "rffr.json"), "w") as f:
    json.dump(pack, f, ensure_ascii=False, indent=2)
    f.write("\n")

# ---- Public website catalogue ----
guideline_order = []
for c in ism:
    if c["guideline"] not in guideline_order:
        guideline_order.append(c["guideline"])

public_controls = [{
    "id": c["id"], "guideline": c["guideline"], "cat": GUIDELINE_CAT.get(c["guideline"], "other"),
    "section": c["section"], "topic": c["topic"], "text": c["desc"],
    "rffrCore": c["rffrCore"],
    "e8": "ML1" if c["ml1"] else "ML2" if c["ml2"] else "ML3" if c["ml3"] else None,
    "ml": {"ml1": c["ml1"], "ml2": c["ml2"], "ml3": c["ml3"]},
    "classification": [x for x in (("NC" if c["nc"] else None), ("OS" if c["os"] else None)) if x],
    "microsoft": ism_to_check.get(c["id"]),
} for c in ism]

automated = sum(1 for c in public_controls if c["microsoft"])
e8n = sum(1 for c in public_controls if c["e8"])
coren = sum(1 for c in public_controls if c["rffrCore"])
# "Microsoft-assessable" = the union of controls with a live Graph
# posture check (directly auto-assessed, one-click) AND the Essential
# Eight-mapped controls (the Microsoft-hardening band the E8 module
# assesses). Honest broader coverage without double counting.
assessable = sum(1 for c in public_controls if c["microsoft"] or c["e8"])

site = {
    "ismVersion": "June 2026",
    "generated": date.today().isoformat(),
    "counts": {
        "ismTotal": len(ism), "deeds": len(deeds), "total": len(ism) + len(deeds),
        "rffrCore": coren, "e8Mapped": e8n, "microsoftAutomated": automated,
        "microsoftAssessable": assessable,
        "guidelines": len(guideline_order),
    },
    "guidelines": [{
        "name": g, "cat": GUIDELINE_CAT.get(g),
        "count": sum(1 for c in public_controls if c["guideline"] == g),
        "automated": sum(1 for c in public_controls if c["guideline"] == g and c["microsoft"]),
        "assessable": sum(1 for c in public_controls if c["guideline"] == g and (c["microsoft"] or c["e8"])),
    } for g in guideline_order],
    "deeds": [{"id": d["id"], "text": d["desc"], "rffrCore": d["rffrCore"]} for d in deeds],
    "microsoftChecks": CHECK_RFFR,
    "controls": public_controls,
}
with open(os.path.join(ROOT, "src", "data", "rffr-controls.json"), "w") as f:
    json.dump(site, f, ensure_ascii=False, indent=2)
    f.write("\n")

print("Wrote checkpoint-content/rffr.json  (%d controls: %d deeds + %d ISM)"
      % (len(pack_controls), len(deeds), len(ism)))
print("Wrote src/data/rffr-controls.json")
print("  RFFR Core Expectations:  %d" % coren)
print("  E8-mapped (ML1/2/3):     %d" % e8n)
print("  Microsoft-automated:     %d (via %d posture checks)" % (automated, len(CHECK_RFFR)))
print("  Guidelines:              %d" % len(guideline_order))
